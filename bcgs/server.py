from flask import Flask, jsonify, send_from_directory, request
import requests
import re
from bs4 import BeautifulSoup
import json
import uuid
from openpyxl import Workbook

from disqus_objects import User
from worksheet_objects import WorkbookManager
from constants import API_KEY, DISQUS_ID_RE, THREAD_DATA_RE

app = Flask(__name__)

#   "author": {
#         ... (see disqus_objects.py)
#     },
#     "canVote": false,
#     "createdAt": "2018-04-05T02:44:32",
#     "dislikes": 0,
#     "forum": "breitbartproduction",
#     "id": "3839799615",
#     "isApproved": true,
#     "isDeleted": false,
#     "isDeletedByAuthor": false,
#     "isEdited": false,
#     "isFlagged": false,
#     "isHighlighted": false,
#     "isSpam": false,
#     "likes": 0,
#     "media": [],
#     "message": "<p>She's just insane.  Another MUSLIM flips out.</p>",
#     "moderationLabels": [],
#     "numReports": 0,
#     "parent": 3839693015,
#     "points": 0,
#     "raw_message": "She's just insane.  Another MUSLIM flips out.",
#     "sb": false,
#     "thread": "6595667472"
# }

_user_labels = [
    'Global ID', 'Display Name', 'Username', 'Total Post Count',
    'Total Like Count', 'Location', 'Join Date', 'Profile Webpage'
]
_comment_labels = [
    'Global ID', 'Author Global Id', 'Source Thread', 'Parent Comment Global ID', 'Sub-thread', 'Date', 'Text', 'Likes'
]
_sheet_info = ['Story Name', 'Story Url']


def add_correct_comment_info(user, thread, comment, comment_ws):
    row = [comment['id'], user.id, thread, comment['parent'], comment['thread']]
    if not user.private:
        row.extend([comment['createdAt'], comment['raw_message'], comment['likes']])

    comment_ws.append(row)


def add_comment(comment, req_info, comment_ws:WorkbookManager, user_ws:WorkbookManager):
    """

    :param dict comment:
    :param dict req_info:
    :param comment_ws:
    :param user_ws:
    :return:
    """
    try:
        user = User(comment['author'])

        add_correct_comment_info(user, req_info['thread'], comment, comment_ws)

        # for private and anonymous users, do not record information
        if user.private:
            return

        if user.id not in req_info['users']:
            # don't double-add users
            req_info['users'].add(user.id)
            user_ws.append(user.user_info_row)
    except:
        print("something went wrong with a comment")
        print(comment)

def add_comments(current_comments, req_info, comment_ws:WorkbookManager, user_ws:WorkbookManager):
    """

    :param dict current_comments:
    :param dict req_info:
    :param comment_ws:
    :param user_ws:
    :return:
    """
    # print(current_comments.keys())
    comments = current_comments['response']

    if type(comments) is dict and 'posts' in comments:
        # first format is different
        comments = comments['posts']
    for comment in comments:
        add_comment(comment, req_info, comment_ws, user_ws)

def get_next_comments(previous_comments, req_info):
    result = None

    if previous_comments['cursor']['hasNext']:
        # get next comments
        # https://disqus.com/api/3.0/threads/listPostsThreaded?limit=100&thread=6595667472&forum=breitbartproduction&order=popular&cursor=1%3A0%3A0&api_key=E8Uh5l5fHZ6gD8U3KycjAIAk46f68Zw7C6eW8WSjZvCLXebZ7p0r1yrYDrLilk2F
        cursor = previous_comments['cursor']['next']
        thread = req_info['thread']
        next_comments = requests.get(
            'https://disqus.com/api/3.0/threads/listPostsThreaded',
            {'limit': 100, 'forum': 'breitbartproduction', 'api_key': API_KEY, 'cursor': cursor, 'thread': thread}
        )
        result = next_comments.json()

    return result


def fetch_article_info(url):
    result = {
        'url': url,
        'title': "Title not found by tool - email this to aeturnum@gmail.com",
        'disqus_id': None
    }

    data = requests.get(url)
    match = re.search(DISQUS_ID_RE, data.text)

    # if the match is wrong, return an empty id
    if match:

        result['disqus_id'] = match.group(1)

        title_parser = BeautifulSoup(data.text, 'html.parser')
        for meta_tag in title_parser.head.find_all('meta'):
            content = meta_tag.get('content', None)
            property = meta_tag.get('property', None)
            if property == 'og:title':
                result['title'] = content

    return result


@app.route('/comments')
def fetch_comments():
    url_arg = request.args.get('URL')
    req_info = {
        'users': set(),
        'thread': None
    }
    # 'http://www.breitbart.com/california/2018/04/04/log_report-nasim-aghdams-brother-warned-police-she-might-do-something/'
    url_info = fetch_article_info(url_arg)

    # first_comments = requests.get(f'https://disqus.com/embed/comments/?base=default&f=breitbartproduction&t_i={disqus_id}')
    first_comments = requests.get(
        'https://disqus.com/embed/comments/',
        params={'base': 'default', 'f': 'breitbartproduction', 't_i': url_info['disqus_id']}
    )

    bs = BeautifulSoup(first_comments.text, 'html.parser')
    comment_json = bs.find(id="disqus-threadData").next
    comment_json = json.loads(comment_json)


    req_info['thread'] = comment_json['response']['thread']['id']

    wb = Workbook()
    main = WorkbookManager(wb.active)
    main.title = "Comment Source"
    main.append(_sheet_info)
    main.append([url_info['title'], url_info['url']])

    comments = WorkbookManager(wb.create_sheet("Comments"))
    comments.append(_comment_labels)
    users = WorkbookManager(wb.create_sheet("Users"))
    users.append(_user_labels)


    while True:
        add_comments(comment_json, req_info, comments, users)
        comment_json = get_next_comments(comment_json, req_info)
        if comment_json is None:
            break


    name = f'{str(uuid.uuid4())}.xlsx'
    wb.save(name)
    return send_from_directory(".", name, attachment_filename="breitbart.xlsx")

@app.route('/')
def index():
    return """
<head></head>
<body>
    <h2>Enter Breitbart comment link</h2>

    <form action="/comments">
      Comment URL:<br>
      <input type="text" size="200" name="URL" value="http://www.breitbart.com/jerusalem/2018/05/09/idf-iranian-forces-fire-rockets-israel/">
      <input type="submit" value="Submit">
    </form>
</body>
    """